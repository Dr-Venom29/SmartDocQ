# backend/tests/test_spreadsheet_preview.py

import io
import os
import sys
import datetime
from unittest.mock import patch

# Set env before imports
os.environ["SERVICE_TOKEN"] = "test_service_token"
os.environ.setdefault("FLASK_DEBUG", "1")

import pytest
from main import app
from openpyxl import Workbook


@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client


def test_spreadsheet_preview_unauthorized(client):
    # No service token
    resp = client.get("/api/document/preview/doc123/spreadsheet")
    assert resp.status_code == 401

    # Invalid service token
    resp = client.get(
        "/api/document/preview/doc123/spreadsheet",
        headers={"x-service-token": "wrong_token"},
    )
    assert resp.status_code == 401


@patch("routes.document_routes.fetch_doc_from_node")
def test_spreadsheet_preview_not_found(mock_fetch, client):
    mock_fetch.return_value = (False, "Document not found", None, None)

    resp = client.get(
        "/api/document/preview/doc123/spreadsheet",
        headers={"x-service-token": "test_service_token"},
    )
    assert resp.status_code == 404
    assert resp.get_json() == {"error": "Document not found"}


@patch("routes.document_routes.fetch_doc_from_node")
def test_spreadsheet_preview_unsupported_type(mock_fetch, client):
    mock_fetch.return_value = (True, "document.pdf", "application/pdf", b"%PDF-1.4...")

    resp = client.get(
        "/api/document/preview/doc123/spreadsheet",
        headers={"x-service-token": "test_service_token"},
    )
    assert resp.status_code == 415
    assert "Unsupported document type" in resp.get_json()["error"]


@patch("routes.document_routes.fetch_doc_from_node")
def test_spreadsheet_preview_csv_success(mock_fetch, client):
    csv_data = b"ID,Name,Score\n1,Rohit,92\n2,Anirudh,88\n"
    mock_fetch.return_value = (True, "grades.csv", "text/csv", csv_data)

    resp = client.get(
        "/api/document/preview/doc123/spreadsheet",
        headers={"x-service-token": "test_service_token"},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["type"] == "csv"
    assert len(data["sheets"]) == 1
    sheet = data["sheets"][0]
    assert sheet["name"] == "CSV"
    assert sheet["headers"] == ["ID", "Name", "Score"]
    assert sheet["rows"] == [["1", "Rohit", "92"], ["2", "Anirudh", "88"]]
    assert sheet["rowCount"] == 2
    assert sheet["columnCount"] == 3
    assert sheet["truncated"] is False


@patch("routes.document_routes.fetch_doc_from_node")
def test_spreadsheet_preview_xlsx_success(mock_fetch, client):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Experiments"
    ws1.append(["ID", "Model", "Accuracy"])
    ws1.append([1, "efficientnet", 0.91])

    ws2 = wb.create_sheet(title="Metrics")
    ws2.append(["Metric", "Value"])
    ws2.append(["Precision", 0.88])

    bio = io.BytesIO()
    wb.save(bio)
    xlsx_bytes = bio.getvalue()

    mock_fetch.return_value = (
        True,
        "results.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        xlsx_bytes,
    )

    resp = client.get(
        "/api/document/preview/doc123/spreadsheet",
        headers={"x-service-token": "test_service_token"},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["type"] == "xlsx"
    assert len(data["sheets"]) == 2

    s1 = data["sheets"][0]
    assert s1["name"] == "Experiments"
    assert s1["headers"] == ["ID", "Model", "Accuracy"]
    assert s1["rows"] == [["1", "efficientnet", "0.91"]]
    assert s1["rowCount"] == 1
    assert s1["columnCount"] == 3
    assert s1["truncated"] is False

    s2 = data["sheets"][1]
    assert s2["name"] == "Metrics"
    assert s2["headers"] == ["Metric", "Value"]
    assert s2["rows"] == [["Precision", "0.88"]]


@patch("routes.document_routes.fetch_doc_from_node")
def test_spreadsheet_preview_limits(mock_fetch, client):
    # Create wide and long CSV
    headers = [f"Col{i}" for i in range(120)]
    rows = [[f"R{r}C{c}" for c in range(120)] for r in range(600)]
    csv_lines = [",".join(headers)] + [",".join(row) for row in rows]
    csv_bytes = "\n".join(csv_lines).encode("utf-8")

    mock_fetch.return_value = (True, "large.csv", "text/csv", csv_bytes)

    resp = client.get(
        "/api/document/preview/doc123/spreadsheet",
        headers={"x-service-token": "test_service_token"},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    sheet = data["sheets"][0]
    assert sheet["rowCount"] == 600
    assert sheet["columnCount"] == 120
    assert sheet["truncated"] is True
    # Verify dimensions are truncated to limits (500 rows, 100 columns)
    assert len(sheet["headers"]) == 100
    assert len(sheet["rows"]) == 500
    assert len(sheet["rows"][0]) == 100


@patch("routes.document_routes.fetch_doc_from_node")
def test_spreadsheet_preview_normalization(mock_fetch, client):
    wb = Workbook()
    ws = wb.active
    ws.title = "Normal"
    # Headers
    ws.append(["Bool", "None", "Date", "DateTime", "String"])
    # Values
    dt = datetime.datetime(2026, 7, 11, 12, 0, 0)
    d = datetime.date(2026, 7, 11)
    ws.append([True, None, d, dt, "Hello"])

    bio = io.BytesIO()
    wb.save(bio)
    xlsx_bytes = bio.getvalue()

    mock_fetch.return_value = (
        True,
        "norm.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        xlsx_bytes,
    )

    resp = client.get(
        "/api/document/preview/doc123/spreadsheet",
        headers={"x-service-token": "test_service_token"},
    )
    assert resp.status_code == 200
    data = resp.get_json()
    sheet = data["sheets"][0]
    row = sheet["rows"][0]
    # Check normalized values
    assert row[0] == "true"
    assert row[1] == ""
    assert row[2] == "2026-07-11 00:00:00"
    assert row[3] == "2026-07-11 12:00:00"
    assert row[4] == "Hello"
